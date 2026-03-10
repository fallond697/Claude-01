"""
Build a Power BI-ready Excel dashboard from ServiceNow Change Management data.
Reads raw change data from a JSON cache file, shapes into pivot tables and summary.
"""
import json
import os
import calendar
from datetime import datetime, timedelta
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter

# ── CONFIG ──
OUTPUT = os.path.expanduser(
    r"~\OneDrive - Vituity\Documents\Change Management\Change_Management_Dashboard_ServiceNow.xlsx"
)
SNOW_URL = "https://vituity.service-now.com/nav_to.do?uri=change_request.do?sysparm_query=number="

# ── LOAD DATA ──
DATA_FILE = os.path.join(os.path.dirname(__file__), "calendar_changes.json")
with open(DATA_FILE, "r", encoding="utf-8") as f:
    all_changes = json.load(f)

# Filter out Canceled, New, and Assess states + limit to last 30 days + upcoming 2 weeks
EXCLUDE_STATES = {"Canceled", "New", "Assess"}
today = datetime.now().date()
window_start = today - timedelta(days=30)
window_end = today + timedelta(days=14)

changes = []
for c in all_changes:
    if c.get("state", "") in EXCLUDE_STATES:
        continue
    ps = c.get("planned_start", "")
    if ps and len(ps) >= 10:
        try:
            dt = datetime.strptime(ps[:10], "%Y-%m-%d").date()
            if window_start <= dt <= window_end:
                changes.append(c)
        except ValueError:
            pass
print(f"Loaded {len(changes)} change records ({window_start} to {window_end}, excl Canceled/New/Assess)")

# ── STYLES ──
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
LINK_FONT = Font(name="Calibri", color="0563C1", underline="single", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC")
)
PCT_FMT = '0.0"%"'
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="003366")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="003366")


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def auto_width(ws, max_width=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def alt_rows(ws, start_row, end_row, col_count):
    for row in range(start_row, end_row + 1):
        if row % 2 == 0:
            for col in range(1, col_count + 1):
                ws.cell(row=row, column=col).fill = ALT_FILL


def write_pivot(ws, title_row, data_dict, col_labels, include_total=True):
    """Write a pivot table. data_dict = {row_label: {col_label: count}}"""
    # Headers
    ws.cell(row=1, column=1, value=title_row)
    for i, label in enumerate(col_labels):
        ws.cell(row=1, column=i + 2, value=label)
    if include_total:
        ws.cell(row=1, column=len(col_labels) + 2, value="Total")

    row_num = 2
    for row_label, counts in data_dict.items():
        ws.cell(row=row_num, column=1, value=row_label)
        row_total = 0
        for i, cl in enumerate(col_labels):
            val = counts.get(cl, 0)
            ws.cell(row=row_num, column=i + 2, value=val)
            row_total += val
        if include_total:
            ws.cell(row=row_num, column=len(col_labels) + 2, value=row_total)
        row_num += 1

    total_cols = len(col_labels) + (2 if include_total else 1)
    style_header(ws, total_cols)
    alt_rows(ws, 2, row_num - 1, total_cols)
    auto_width(ws)
    return row_num


# ── PARSE FIELDS ──
CHANGE_TYPES = ["Normal", "Standard", "Emergency"]

for c in changes:
    # Parse month from planned start
    psd = c.get("planned_start", "") or c.get("start_date", "") or ""
    if psd:
        try:
            dt = datetime.strptime(psd[:10], "%Y-%m-%d")
            c["_month"] = dt.strftime("%Y-%m")
            c["_date"] = dt
        except (ValueError, TypeError):
            c["_month"] = "Unknown"
            c["_date"] = None
    else:
        c["_month"] = "Unknown"
        c["_date"] = None

    # Normalize type
    t = (c.get("type", "") or "").strip()
    if t.lower() == "normal":
        c["_type"] = "Normal"
    elif t.lower() == "standard":
        c["_type"] = "Standard"
    elif t.lower() == "emergency":
        c["_type"] = "Emergency"
    else:
        c["_type"] = t or "Unknown"

    # Normalize state
    c["_state"] = (c.get("state", "") or "").strip()

    # Close code
    c["_close_code"] = (c.get("close_code", "") or "").strip()


# ── CREATE WORKBOOK ──
wb = Workbook()

# ─── Sheet 1: Raw Data ───
ws_raw = wb.active
ws_raw.title = "Raw Data"

raw_headers = [
    "Number", "Short Description", "Type", "State", "Assignment Group",
    "Assigned To", "Planned Start", "Planned End", "Configuration Item",
    "Environment", "Close Code", "Risk"
]
for i, h in enumerate(raw_headers, 1):
    ws_raw.cell(row=1, column=i, value=h)

for row_idx, c in enumerate(changes, 2):
    num = c.get("number", "")
    ws_raw.cell(row=row_idx, column=1, value=num)
    if num.startswith("CHG"):
        ws_raw.cell(row=row_idx, column=1).hyperlink = SNOW_URL + num
        ws_raw.cell(row=row_idx, column=1).font = LINK_FONT

    ws_raw.cell(row=row_idx, column=2, value=c.get("short_description", ""))
    ws_raw.cell(row=row_idx, column=3, value=c["_type"])
    ws_raw.cell(row=row_idx, column=4, value=c["_state"])
    ws_raw.cell(row=row_idx, column=5, value=c.get("assignment_group", ""))
    ws_raw.cell(row=row_idx, column=6, value=c.get("assigned_to", ""))
    ws_raw.cell(row=row_idx, column=7, value=c.get("planned_start", c.get("start_date", "")))
    ws_raw.cell(row=row_idx, column=8, value=c.get("planned_end", c.get("end_date", "")))
    ws_raw.cell(row=row_idx, column=9, value=c.get("cmdb_ci", c.get("configuration_item", "")))
    ws_raw.cell(row=row_idx, column=10, value=c.get("environment", ""))
    ws_raw.cell(row=row_idx, column=11, value=c["_close_code"])
    ws_raw.cell(row=row_idx, column=12, value=c.get("risk", ""))

style_header(ws_raw, len(raw_headers))
alt_rows(ws_raw, 2, len(changes) + 1, len(raw_headers))
auto_width(ws_raw)

# ─── Sheet 2: Changes by Month ───
ws_month = wb.create_sheet("Changes by Month")
month_data = defaultdict(lambda: defaultdict(int))
for c in changes:
    month_data[c["_month"]][c["_type"]] += 1
# Sort by month
month_sorted = dict(sorted(month_data.items()))
write_pivot(ws_month, "Month", month_sorted, CHANGE_TYPES)

# ─── Sheet 3: Changes by Config Item ───
ws_ci = wb.create_sheet("Changes by Config Item")
ci_data = defaultdict(lambda: defaultdict(int))
ci_totals = Counter()
for c in changes:
    ci = c.get("cmdb_ci", c.get("configuration_item", "")) or "(not set)"
    ci_data[ci][c["_type"]] += 1
    ci_totals[ci] += 1
ci_sorted = {k: ci_data[k] for k, _ in ci_totals.most_common(20)}
write_pivot(ws_ci, "Configuration Item", ci_sorted, CHANGE_TYPES)

# ─── Sheet 4: Changes by State ───
ws_state = wb.create_sheet("Changes by State")
state_data = defaultdict(lambda: defaultdict(int))
for c in changes:
    state_data[c["_state"]][c["_type"]] += 1
write_pivot(ws_state, "State", dict(sorted(state_data.items())), CHANGE_TYPES)

# ─── Sheet 5: Changes by Assignment Group ───
ws_ag = wb.create_sheet("Changes by Assignment Group")
ag_data = defaultdict(lambda: defaultdict(int))
ag_totals = Counter()
for c in changes:
    ag = c.get("assignment_group", "") or "(not set)"
    ag_data[ag][c["_type"]] += 1
    ag_totals[ag] += 1
ag_sorted = {k: ag_data[k] for k, _ in ag_totals.most_common(15)}
write_pivot(ws_ag, "Assignment Group", ag_sorted, CHANGE_TYPES)

# ─── Sheet 6: Changes by Environment ───
ws_env = wb.create_sheet("Changes by Environment")
env_data = defaultdict(lambda: defaultdict(int))
for c in changes:
    env = c.get("environment", "") or "(not set)"
    env_data[env][c["_type"]] += 1
write_pivot(ws_env, "Environment", dict(sorted(env_data.items())), CHANGE_TYPES)

# ─── Sheet 7: Monthly Trend ───
ws_trend = wb.create_sheet("Monthly Trend")
trend_headers = ["Month", "Total", "Normal", "Standard", "Emergency", "Closed", "Successful", "Success Rate"]
for i, h in enumerate(trend_headers, 1):
    ws_trend.cell(row=1, column=i, value=h)

row_num = 2
for month in sorted(month_data.keys()):
    counts = month_data[month]
    total = sum(counts.values())
    # Count closed and successful for this month
    closed = sum(1 for c in changes if c["_month"] == month and c["_state"] == "Closed")
    successful = sum(1 for c in changes if c["_month"] == month and c["_close_code"] == "Successful")
    rate = (successful / closed * 100) if closed > 0 else 0

    ws_trend.cell(row=row_num, column=1, value=month)
    ws_trend.cell(row=row_num, column=2, value=total)
    ws_trend.cell(row=row_num, column=3, value=counts.get("Normal", 0))
    ws_trend.cell(row=row_num, column=4, value=counts.get("Standard", 0))
    ws_trend.cell(row=row_num, column=5, value=counts.get("Emergency", 0))
    ws_trend.cell(row=row_num, column=6, value=closed)
    ws_trend.cell(row=row_num, column=7, value=successful)
    ws_trend.cell(row=row_num, column=8, value=round(rate, 1))
    ws_trend.cell(row=row_num, column=8).number_format = '0.0"%"'
    row_num += 1

style_header(ws_trend, len(trend_headers))
alt_rows(ws_trend, 2, row_num - 1, len(trend_headers))
auto_width(ws_trend)

# ─── Sheet 8: Dashboard Summary ───
ws_dash = wb.create_sheet("Dashboard Summary")

total_changes = len(changes)
type_counts = Counter(c["_type"] for c in changes)
state_counts = Counter(c["_state"] for c in changes)
months_with_data = len([m for m in month_data if m != "Unknown"])
avg_per_month = round(total_changes / months_with_data, 1) if months_with_data else 0
total_closed = state_counts.get("Closed", 0)
total_successful = sum(1 for c in changes if c["_close_code"] == "Successful")
success_rate = round(total_successful / total_closed * 100, 1) if total_closed > 0 else 0

row = 1
ws_dash.cell(row=row, column=1, value="Change Management Dashboard").font = TITLE_FONT
ws_dash.cell(row=row, column=3, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = NORMAL_FONT
row += 1
ws_dash.cell(row=row, column=1, value="Source: ServiceNow (Live Query)").font = NORMAL_FONT
row += 2

# Key Metrics
ws_dash.cell(row=row, column=1, value="KEY METRICS").font = SUBTITLE_FONT
row += 1
metrics = [
    ("Total Changes", total_changes),
    ("Average per Month", avg_per_month),
    ("Overall Success Rate", f"{success_rate}%"),
    ("Total Closed", total_closed),
    ("Total Canceled", state_counts.get("Canceled", 0)),
]
for label, val in metrics:
    ws_dash.cell(row=row, column=1, value=label).font = BOLD_FONT
    ws_dash.cell(row=row, column=2, value=val).font = NORMAL_FONT
    row += 1

row += 1
ws_dash.cell(row=row, column=1, value="BY TYPE").font = SUBTITLE_FONT
row += 1
for t in CHANGE_TYPES:
    ws_dash.cell(row=row, column=1, value=t).font = BOLD_FONT
    ws_dash.cell(row=row, column=2, value=type_counts.get(t, 0)).font = NORMAL_FONT
    row += 1

row += 1
ws_dash.cell(row=row, column=1, value="BY STATE").font = SUBTITLE_FONT
row += 1
for state, count in state_counts.most_common():
    ws_dash.cell(row=row, column=1, value=state).font = BOLD_FONT
    ws_dash.cell(row=row, column=2, value=count).font = NORMAL_FONT
    row += 1

row += 1
ws_dash.cell(row=row, column=1, value="TOP 10 ASSIGNMENT GROUPS").font = SUBTITLE_FONT
row += 1
for ag, count in ag_totals.most_common(10):
    ws_dash.cell(row=row, column=1, value=ag).font = NORMAL_FONT
    ws_dash.cell(row=row, column=2, value=count).font = NORMAL_FONT
    row += 1

row += 1
ws_dash.cell(row=row, column=1, value="TOP 10 CONFIGURATION ITEMS").font = SUBTITLE_FONT
row += 1
for ci, count in ci_totals.most_common(10):
    ws_dash.cell(row=row, column=1, value=ci).font = NORMAL_FONT
    ws_dash.cell(row=row, column=2, value=count).font = NORMAL_FONT
    row += 1

ws_dash.column_dimensions["A"].width = 35
ws_dash.column_dimensions["B"].width = 15
ws_dash.column_dimensions["C"].width = 25

# ─── Sheet 9: Full Calendar View ───
# Mimics ServiceNow's "View Full Calendar" — monthly grids with changes in day cells
ws_cal = wb.create_sheet("Full Calendar")

# Color fills for change types
TYPE_FILLS = {
    "Normal": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),     # Blue
    "Standard": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),    # Green
    "Emergency": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),   # Red
}
TYPE_FONTS = {
    "Normal": Font(name="Calibri", size=9, color="FFFFFF"),
    "Standard": Font(name="Calibri", size=9, color="FFFFFF"),
    "Emergency": Font(name="Calibri", size=9, color="FFFFFF", bold=True),
}
DAY_NUM_FONT = Font(name="Calibri", size=11, bold=True, color="333333")
DAY_NUM_FONT_GRAY = Font(name="Calibri", size=11, bold=True, color="BBBBBB")
MONTH_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="003366")
DOW_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
DOW_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
WEEKEND_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
TODAY_BORDER = Border(
    left=Side(style="medium", color="FF6600"),
    right=Side(style="medium", color="FF6600"),
    top=Side(style="medium", color="FF6600"),
    bottom=Side(style="medium", color="FF6600"),
)
CELL_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
LEGEND_FILL_BG = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")

# Build a dict: date_str -> list of changes
changes_by_date = defaultdict(list)
for c in changes:
    ps = c.get("planned_start", "")
    if ps and len(ps) >= 10:
        date_key = ps[:10]  # YYYY-MM-DD
        changes_by_date[date_key].append(c)

# Determine month range from data
all_months = set()
for c in changes:
    ps = c.get("planned_start", "")
    if ps and len(ps) >= 7:
        all_months.add(ps[:7])  # YYYY-MM

all_months = sorted(all_months)
if not all_months:
    all_months = [datetime.now().strftime("%Y-%m")]

# Calendar grid settings
COL_WIDTH = 22       # Width of each day column
ROW_HEIGHT_DAY = 15  # Height for day number row
ROW_HEIGHT_CHG = 13  # Height per change entry row
MAX_CHANGES_PER_DAY = 6  # Max visible entries per day cell
ROWS_PER_WEEK = 1 + MAX_CHANGES_PER_DAY  # day number + change rows
DOW_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

# Set column widths (cols A-G for the 7 days)
for col_idx in range(1, 8):
    ws_cal.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTH

current_row = 1
today = datetime.now().date()

# Legend at top
ws_cal.cell(row=current_row, column=1, value="Change Calendar").font = Font(name="Calibri", size=16, bold=True, color="003366")
current_row += 1
ws_cal.cell(row=current_row, column=1, value="Source: ServiceNow | Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M")).font = Font(name="Calibri", size=9, color="666666")
current_row += 1

# Legend row
for i, (tname, tfill) in enumerate(TYPE_FILLS.items()):
    col = (i * 2) + 1
    cell = ws_cal.cell(row=current_row, column=col, value=f"  {tname}  ")
    cell.fill = tfill
    cell.font = TYPE_FONTS[tname]
    cell.alignment = Alignment(horizontal="center")
current_row += 2

# Render each month
for month_str in all_months:
    year, month = int(month_str[:4]), int(month_str[5:7])
    month_name = calendar.month_name[month]

    # Month title
    ws_cal.cell(row=current_row, column=1, value=f"{month_name} {year}").font = MONTH_TITLE_FONT
    ws_cal.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    current_row += 1

    # Day-of-week headers
    for col_idx, dow in enumerate(DOW_NAMES, 1):
        cell = ws_cal.cell(row=current_row, column=col_idx, value=dow)
        cell.font = DOW_FONT
        cell.fill = DOW_FILL
        cell.alignment = Alignment(horizontal="center")
    current_row += 1

    # Get the calendar grid for this month (Sunday-start)
    cal = calendar.Calendar(firstweekday=6)  # Sunday start
    month_days = cal.monthdayscalendar(year, month)

    for week in month_days:
        week_start_row = current_row

        # Day number row
        for col_idx, day in enumerate(week, 1):
            cell = ws_cal.cell(row=current_row, column=col_idx)
            if day != 0:
                cell.value = day
                cell.font = DAY_NUM_FONT
                cell.alignment = Alignment(horizontal="left", vertical="top")

                # Highlight today
                try:
                    this_date = datetime(year, month, day).date()
                    if this_date == today:
                        cell.border = TODAY_BORDER
                except ValueError:
                    pass

                # Weekend shading
                if col_idx in (1, 7):  # Sunday, Saturday
                    cell.fill = WEEKEND_FILL
            else:
                cell.font = DAY_NUM_FONT_GRAY
                cell.fill = WEEKEND_FILL

            cell.border = CELL_BORDER

        current_row += 1

        # Change entry rows for this week
        for entry_idx in range(MAX_CHANGES_PER_DAY):
            for col_idx, day in enumerate(week, 1):
                cell = ws_cal.cell(row=current_row, column=col_idx)
                cell.border = CELL_BORDER

                if day != 0:
                    date_key = f"{year:04d}-{month:02d}-{day:02d}"
                    day_changes = changes_by_date.get(date_key, [])

                    if col_idx in (1, 7):
                        cell.fill = WEEKEND_FILL

                    if entry_idx < len(day_changes):
                        chg = day_changes[entry_idx]
                        chg_num = chg.get("number", "")
                        short_desc = chg.get("short_description", "")
                        chg_type = chg.get("type", "Normal")

                        # Truncate description to fit
                        label = f"{chg_num}: {short_desc}"
                        if len(label) > 30:
                            label = label[:28] + ".."

                        cell.value = label
                        cell.fill = TYPE_FILLS.get(chg_type, TYPE_FILLS["Normal"])
                        cell.font = TYPE_FONTS.get(chg_type, TYPE_FONTS["Normal"])
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

                        # Add hyperlink
                        if chg_num.startswith("CHG"):
                            cell.hyperlink = SNOW_URL + chg_num

                    elif entry_idx == MAX_CHANGES_PER_DAY - 1 and len(day_changes) > MAX_CHANGES_PER_DAY:
                        remaining = len(day_changes) - MAX_CHANGES_PER_DAY + 1
                        cell.value = f"  +{remaining} more..."
                        cell.font = Font(name="Calibri", size=8, italic=True, color="666666")
                else:
                    cell.fill = WEEKEND_FILL

            ws_cal.row_dimensions[current_row].height = ROW_HEIGHT_CHG
            current_row += 1

        # Set day number row height
        ws_cal.row_dimensions[week_start_row].height = ROW_HEIGHT_DAY

    # Gap between months
    current_row += 2

print(f"  Calendar view: {len(all_months)} months rendered")

# ── SAVE ──
wb.save(OUTPUT)
print(f"\nDashboard saved to: {OUTPUT}")
print(f"  {len(changes)} changes across {months_with_data} months")
print(f"  Types: {dict(type_counts)}")
print(f"  States: {dict(state_counts)}")
print(f"  Success rate: {success_rate}%")
print(f"  9 sheets: Raw Data, Changes by Month, Config Item, State, Assignment Group, Environment, Monthly Trend, Dashboard Summary, Full Calendar")
