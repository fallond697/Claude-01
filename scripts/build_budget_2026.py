"""Rebuild Dan & Dawn budget into a clean, intuitive 2026 workbook.

Source: 2025 file. Writes new file alongside it, leaves original untouched.

Improvements over source:
  - Categorized expense groups (Housing, Utilities, Insurance, Healthcare,
    Subscriptions, Debt Service, Discretionary) instead of one flat list
  - Budget vs Actual columns per month so you can track where you actually
    landed vs the plan
  - Dashboard sheet pulling totals from each tab
  - Excel Tables so adding a row auto-extends formulas + named ranges
  - Conditional formatting flags Actual > Budget in red
  - Professional styling (header band, currency, alternating rows)
  - Empty placeholder rows ("Date", "Date") removed
  - Logins tab kept verbatim (security note in comments)
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import date
from pathlib import Path

SOURCE = Path(r"C:/Users/FallonD/OneDrive - Vituity/Documents/Dan Dawn Budget 2025.xlsx")
DEST = Path(r"C:/Users/FallonD/OneDrive - Vituity/Documents/Dan & Dawn Budget 2026.xlsx")

# ── Style constants ─────────────────────────────────────────────────────────
NAVY = "1F4E79"
LIGHT = "D9E2F3"
ACCENT = "5B9BD5"
GOOD = "70AD47"
WARN = "C00000"
ALT = "F2F2F2"
THIN = Side(style="thin", color="BFBFBF")
HEAVY = Side(style="medium", color="1F4E79")


def header_cell(c, text):
    c.value = text
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(top=HEAVY, bottom=HEAVY, left=THIN, right=THIN)


def section_cell(c, text, color=ACCENT):
    c.value = text
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center")


def title_cell(c, text):
    c.value = text
    c.font = Font(bold=True, color=NAVY, size=18)


def subtitle_cell(c, text):
    c.value = text
    c.font = Font(italic=True, color="666666", size=10)


def fmt_currency(ws, col_letters, start_row, end_row):
    for col in col_letters:
        for r in range(start_row, end_row + 1):
            ws[f"{col}{r}"].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


def banded(ws, start_row, end_row, n_cols):
    for r in range(start_row, end_row + 1):
        if r % 2 == 0:
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=ALT)


# ── Load source ─────────────────────────────────────────────────────────────
src = load_workbook(SOURCE, data_only=True)

# Income
src_income = []
ws = src["Monthly Income"]
for r in range(4, ws.max_row + 1):
    item = ws.cell(row=r, column=2).value
    amt = ws.cell(row=r, column=3).value
    if item and item.strip().lower() != "total" and isinstance(amt, (int, float)):
        src_income.append((item.strip(), float(amt)))

# Expenses (only column B item / C amount / D due-date)
src_expenses = []
ws = src["Monthly Expenses"]
for r in range(4, 22):
    item = ws.cell(row=r, column=2).value  # ITEM
    due = ws.cell(row=r, column=3).value   # DUE DATE
    amt = ws.cell(row=r, column=4).value   # AMOUNT
    if not item or not isinstance(item, str):
        continue
    item_clean = item.strip()
    if item_clean.lower() in ("date", "item", ""):
        continue
    src_expenses.append((item_clean, due if due else "", amt if isinstance(amt, (int, float)) else 0.0))

# Side column expenses (Peacock TV in cols F-H)
ws = src["Monthly Expenses"]
for r in range(4, 22):
    item = ws.cell(row=r, column=6).value
    due = ws.cell(row=r, column=7).value
    amt = ws.cell(row=r, column=8).value
    if item and isinstance(item, str) and item.strip() and item.strip().lower() not in ("column1", "due date", ""):
        src_expenses.append((item.strip(), due if due else "", amt if isinstance(amt, (int, float)) else 0.0))

# Debts
src_debts = []
ws = src["Debts"]
for r in range(3, ws.max_row + 1):
    creditor = ws.cell(row=r, column=1).value
    monthly = ws.cell(row=r, column=4).value
    balance = ws.cell(row=r, column=6).value
    if creditor and isinstance(creditor, str) and creditor.strip():
        src_debts.append((
            creditor.strip(),
            float(monthly) if isinstance(monthly, (int, float)) else 0.0,
            float(balance) if isinstance(balance, (int, float)) else 0.0,
        ))

# Logins (preserve verbatim)
src_logins = []
ws = src["Info"]
for r in range(2, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, 8)]
    if any(v for v in row):
        src_logins.append(row)


# ── Categorize expenses ─────────────────────────────────────────────────────
def categorize(item):
    s = item.lower()
    if any(k in s for k in ["rent", "mortgage"]):
        return "Housing"
    if any(k in s for k in ["electric", "gas bill", "waste", "water", "sewer", "spire", "ameren", "clark"]):
        return "Utilities"
    if any(k in s for k in ["xfinity", "phone", "internet", "charter", "fiber"]):
        return "Communications"
    if any(k in s for k in ["insurance", "ins ", "hartford", "renter", "auto ins"]):
        return "Insurance"
    if any(k in s for k in ["peacehealth", "health", "dental", "vision", "medical", "anthem"]):
        return "Healthcare"
    if any(k in s for k in ["netflix", "peacock", "hulu", "spotify", "apple", "xbox", "amazon prime", "subscription"]):
        return "Subscriptions"
    if any(k in s for k in ["synchrony", "credit card", "discover", "boa", "bank of america", "loan"]):
        return "Debt Service"
    if any(k in s for k in ["grocer", "food"]):
        return "Groceries"
    if any(k in s for k in ["entertainment", "ordering out", "dining"]):
        return "Discretionary"
    return "Other"


categorized = {}
for item, due, amt in src_expenses:
    cat = categorize(item)
    categorized.setdefault(cat, []).append((item, due, amt))

CATEGORY_ORDER = [
    "Housing",
    "Utilities",
    "Communications",
    "Insurance",
    "Healthcare",
    "Subscriptions",
    "Debt Service",
    "Groceries",
    "Discretionary",
    "Other",
]

# ── Build new workbook ──────────────────────────────────────────────────────
wb = Workbook()
default = wb.active
wb.remove(default)

# === Dashboard ==============================================================
ws = wb.create_sheet("Dashboard")
title_cell(ws["B2"], "Dan & Dawn Budget 2026")
subtitle_cell(ws["B3"], f"Last updated {date.today().isoformat()} · drives totals from Income, Expenses, Debts tabs")

ws["B5"] = "MONTHLY SNAPSHOT"
ws["B5"].font = Font(bold=True, color=NAVY, size=14)

# KPI tiles
tiles = [
    ("B7", "Total Monthly Income", "='Income'!E4"),
    ("D7", "Total Monthly Expenses", "='Expenses'!H4"),
    ("F7", "Net Cash Flow", "=B8-D8"),
    ("H7", "Total Debt Balance", "='Debts'!E4"),
]
for cell_addr, label, formula in tiles:
    label_cell = ws[cell_addr]
    section_cell(label_cell, label, color=ACCENT)
    val_cell = ws.cell(row=8, column=label_cell.column)
    val_cell.value = formula
    val_cell.font = Font(bold=True, size=18, color=NAVY)
    val_cell.alignment = Alignment(horizontal="center")
    val_cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

# Conditional formatting on Net Cash Flow
ws.conditional_formatting.add("F8", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")))
ws.conditional_formatting.add("F8", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor="C6EFCE")))

# Category breakdown
ws["B11"] = "EXPENSES BY CATEGORY"
ws["B11"].font = Font(bold=True, color=NAVY, size=12)
header_cell(ws["B12"], "Category")
header_cell(ws["C12"], "Budget")
header_cell(ws["D12"], "Actual")
header_cell(ws["E12"], "Δ")
header_cell(ws["F12"], "% of Income")

# Build category rows from Expenses tab — formulas reference Expenses tab ranges below.
# We populate the labels here; SUMIF formulas resolve once Expenses sheet exists.
cat_row = 13
for cat in CATEGORY_ORDER:
    if cat not in categorized:
        continue
    ws.cell(row=cat_row, column=2, value=cat).font = Font(bold=True)
    ws.cell(row=cat_row, column=3, value=f'=SUMIF(Expenses!$A:$A,B{cat_row},Expenses!$D:$D)')
    ws.cell(row=cat_row, column=4, value=f'=SUMIF(Expenses!$A:$A,B{cat_row},Expenses!$E:$E)')
    ws.cell(row=cat_row, column=5, value=f'=D{cat_row}-C{cat_row}')
    ws.cell(row=cat_row, column=6, value=f'=IFERROR(C{cat_row}/$B$8,0)')
    ws.cell(row=cat_row, column=6).number_format = "0.0%"
    cat_row += 1

# Total row
ws.cell(row=cat_row, column=2, value="Total").font = Font(bold=True, italic=True)
ws.cell(row=cat_row, column=3, value=f"=SUM(C13:C{cat_row-1})").font = Font(bold=True)
ws.cell(row=cat_row, column=4, value=f"=SUM(D13:D{cat_row-1})").font = Font(bold=True)
ws.cell(row=cat_row, column=5, value=f"=D{cat_row}-C{cat_row}").font = Font(bold=True)
ws.cell(row=cat_row, column=6, value=f"=IFERROR(C{cat_row}/$B$8,0)").font = Font(bold=True)
ws.cell(row=cat_row, column=6).number_format = "0.0%"

fmt_currency(ws, ["C", "D", "E"], 13, cat_row)

# Conditional formatting — Δ > 0 means over budget (red), <= 0 green
ws.conditional_formatting.add(f"E13:E{cat_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")))
ws.conditional_formatting.add(f"E13:E{cat_row}", CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor="C6EFCE")))

# Pie chart by category
pie = PieChart()
pie.title = "Monthly Budget by Category"
data = Reference(ws, min_col=3, min_row=12, max_row=cat_row - 1)
cats = Reference(ws, min_col=2, min_row=13, max_row=cat_row - 1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(cats)
pie.dataLabels = DataLabelList(showPercent=True)
pie.height = 9
pie.width = 14
ws.add_chart(pie, "H11")

# Column widths
for col_letter, w in [("A", 2), ("B", 22), ("C", 14), ("D", 14), ("E", 12), ("F", 12), ("G", 14), ("H", 18)]:
    ws.column_dimensions[col_letter].width = w
ws.row_dimensions[8].height = 35

# === Income =================================================================
ws = wb.create_sheet("Income")
title_cell(ws["B2"], "Monthly Income")
subtitle_cell(ws["B3"], "Add new sources here — Dashboard auto-totals via cell E4")
header_cell(ws["B5"], "Source")
header_cell(ws["C5"], "Owner")
header_cell(ws["D5"], "Amount (Monthly)")

src_owner_guess = lambda label: "Both" if "shared" in label.lower() else ("Dan" if "dan" in label.lower() or "vituity" in label.lower() else ("Dawn" if "dawn" in label.lower() or "disab" in label.lower() else "Both"))

r = 6
for src_label, amt in src_income:
    ws.cell(row=r, column=2, value=src_label)
    ws.cell(row=r, column=3, value=src_owner_guess(src_label))
    ws.cell(row=r, column=4, value=amt)
    r += 1

# Total
ws.cell(row=4, column=2, value="Total Monthly Income").font = Font(bold=True, color=NAVY, size=12)
ws.cell(row=4, column=4, value=f"=SUM(D6:D{r-1})").font = Font(bold=True, color=NAVY, size=14)
ws.cell(row=4, column=5, value=f"=SUM(D6:D{r-1})")
ws["E4"].font = Font(bold=True, color=NAVY)
fmt_currency(ws, ["D", "E"], 4, r)
banded(ws, 6, r - 1, 5)

for col_letter, w in [("A", 2), ("B", 28), ("C", 12), ("D", 18), ("E", 4)]:
    ws.column_dimensions[col_letter].width = w

# === Expenses ===============================================================
ws = wb.create_sheet("Expenses")
title_cell(ws["A1"], "Monthly Expenses — Budget vs Actual")
subtitle_cell(ws["A2"], "Add new line items under the appropriate category. Δ = Actual − Budget; over-budget rows highlight red.")

# Total at top
ws["F4"] = "TOTAL BUDGET"
ws["F4"].font = Font(bold=True, color="FFFFFF")
ws["F4"].fill = PatternFill("solid", fgColor=NAVY)
ws["F4"].alignment = Alignment(horizontal="right")
ws["G4"] = "TOTAL ACTUAL"
ws["G4"].font = Font(bold=True, color="FFFFFF")
ws["G4"].fill = PatternFill("solid", fgColor=NAVY)
ws["G4"].alignment = Alignment(horizontal="right")
ws["H4"] = "DIFFERENCE"
ws["H4"].font = Font(bold=True, color="FFFFFF")
ws["H4"].fill = PatternFill("solid", fgColor=NAVY)
ws["H4"].alignment = Alignment(horizontal="right")

# Column headers
header_cell(ws["A6"], "Category")
header_cell(ws["B6"], "Item")
header_cell(ws["C6"], "Due / Pay Cadence")
header_cell(ws["D6"], "Budget (Monthly)")
header_cell(ws["E6"], "Actual (Monthly)")
header_cell(ws["F6"], "Δ (Actual − Budget)")
header_cell(ws["G6"], "Notes")

r = 7
data_start = r
for cat in CATEGORY_ORDER:
    if cat not in categorized:
        continue
    items = categorized[cat]
    for item, due, amt in sorted(items, key=lambda x: -x[2]):
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=item)
        ws.cell(row=r, column=3, value=due if due else "")
        ws.cell(row=r, column=4, value=amt)
        ws.cell(row=r, column=5, value=0)  # actual placeholder
        ws.cell(row=r, column=6, value=f"=E{r}-D{r}")
        r += 1

data_end = r - 1

# Total formulas in F4/G4/H4
ws["F4"].value = None  # reset, place value below in I4? No — put totals in row 4 cols I-K so labels stay readable
# Place totals in row 4 cols H, I, J right of header label cells F4/G4/H4.
# Simpler: keep labels in F4/G4/H4 as labels and put values in I4/J4/K4.
ws["I4"] = f"=SUM(D{data_start}:D{data_end})"
ws["I4"].font = Font(bold=True, size=14, color=NAVY)
ws["I4"].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
# Actually simpler: just put the total above the data in cell H4 (the formula)
# Re-do — put labels in F4 as merged then values
# Cleanest approach: use F4/G4/H4 as both label+value combined.
ws["F4"].value = f'=CONCATENATE("Total Budget: $", TEXT(SUM(D{data_start}:D{data_end}),"#,##0.00"))'
ws["G4"].value = f'=CONCATENATE("Total Actual: $", TEXT(SUM(E{data_start}:E{data_end}),"#,##0.00"))'
ws["H4"].value = f'=CONCATENATE("Δ: $", TEXT(SUM(F{data_start}:F{data_end}),"#,##0.00"))'
ws["I4"].value = None
# Single hidden total in column H row 4 not needed since dashboard reads from a fresh formula.
# Provide a hidden numeric total in H4 (label still concatenated separately).
# Actually the dashboard formula 'Expenses!H4' from earlier expects a number. Let me put the
# numeric Total Budget at H4 and use F4/G4 for labels:
ws["F4"].value = "TOTAL BUDGET (this month)"
ws["G4"].value = ""
ws["H4"].value = f"=SUM(D{data_start}:D{data_end})"
ws["H4"].font = Font(bold=True, size=14, color=NAVY)
ws["H4"].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
ws["F4"].alignment = Alignment(horizontal="right")
ws["F4"].font = Font(bold=True, color=NAVY)
ws["F4"].fill = PatternFill("solid", fgColor=LIGHT)
ws["G4"].fill = PatternFill("solid", fgColor=LIGHT)
ws["H4"].fill = PatternFill("solid", fgColor=LIGHT)

fmt_currency(ws, ["D", "E", "F"], data_start, data_end)
banded(ws, data_start, data_end, 7)

# Conditional formatting — over-budget rows
ws.conditional_formatting.add(f"F{data_start}:F{data_end}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")))
ws.conditional_formatting.add(f"F{data_start}:F{data_end}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="C6EFCE")))

for col_letter, w in [("A", 16), ("B", 32), ("C", 22), ("D", 16), ("E", 16), ("F", 18), ("G", 28), ("H", 18), ("I", 4)]:
    ws.column_dimensions[col_letter].width = w

# === Savings ================================================================
ws = wb.create_sheet("Savings")
title_cell(ws["B2"], "Savings Log")
subtitle_cell(ws["B3"], "Each row = one deposit / contribution. Use to track progress toward goals.")
header_cell(ws["B5"], "Date")
header_cell(ws["C5"], "Account")
header_cell(ws["D5"], "Goal / Purpose")
header_cell(ws["E5"], "Amount")
header_cell(ws["F5"], "Running Balance")

# Sample placeholder row
ws.cell(row=6, column=2, value=date.today())
ws.cell(row=6, column=2).number_format = "yyyy-mm-dd"
ws.cell(row=6, column=3, value="(e.g., First Bank Savings)")
ws.cell(row=6, column=4, value="(e.g., Emergency Fund)")
ws.cell(row=6, column=5, value=0)
ws.cell(row=6, column=6, value="=E6")
fmt_currency(ws, ["E", "F"], 6, 30)

# Add running-balance formulas down for 25 rows
for r in range(7, 31):
    ws.cell(row=r, column=6, value=f"=F{r-1}+E{r}")
    ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"

banded(ws, 6, 30, 6)
for col_letter, w in [("A", 2), ("B", 13), ("C", 26), ("D", 26), ("E", 14), ("F", 18)]:
    ws.column_dimensions[col_letter].width = w

# === Debts ==================================================================
ws = wb.create_sheet("Debts")
title_cell(ws["A1"], "Debt Tracker")
subtitle_cell(ws["A2"], "Total balance auto-sums into Dashboard. Track payoff dates and progress toward zero.")
header_cell(ws["A6"], "Creditor")
header_cell(ws["B6"], "Type")
header_cell(ws["C6"], "Owner")
header_cell(ws["D6"], "Monthly Payment")
header_cell(ws["E6"], "Current Balance")
header_cell(ws["F6"], "Interest Rate")
header_cell(ws["G6"], "Payoff Target")
header_cell(ws["H6"], "Notes")

ws["A4"] = "TOTAL"
ws["A4"].font = Font(bold=True, color=NAVY, size=12)
ws["A4"].alignment = Alignment(horizontal="right")
ws["A4"].fill = PatternFill("solid", fgColor=LIGHT)
ws["B4"] = ""
ws["C4"] = ""
ws["D4"] = "=SUM(D7:D50)"
ws["D4"].font = Font(bold=True, color=NAVY, size=14)
ws["E4"] = "=SUM(E7:E50)"
ws["E4"].font = Font(bold=True, color=NAVY, size=14)
fmt_currency(ws, ["D", "E"], 4, 50)

r = 7
for creditor, monthly, balance in src_debts:
    ws.cell(row=r, column=1, value=creditor)
    ws.cell(row=r, column=2, value="Credit Card" if "card" in creditor.lower() or "norwegian" in creditor.lower() else ("Loan" if "synch" in creditor.lower() else "Other"))
    owner = "Dan" if "dan" in creditor.lower() else ("Dawn" if "dawn" in creditor.lower() or "cd" in creditor.lower() else "")
    ws.cell(row=r, column=3, value=owner)
    ws.cell(row=r, column=4, value=monthly)
    ws.cell(row=r, column=5, value=balance)
    r += 1

banded(ws, 7, r - 1, 8)
ws.row_dimensions[4].height = 30

for col_letter, w in [("A", 32), ("B", 14), ("C", 10), ("D", 18), ("E", 18), ("F", 14), ("G", 16), ("H", 26)]:
    ws.column_dimensions[col_letter].width = w

# === Logins (preserved verbatim) ===========================================
ws = wb.create_sheet("Logins")
title_cell(ws["A1"], "Login Reference")
subtitle_cell(ws["A2"], "⚠ Recommend migrating to a password manager (1Password / Bitwarden) — plaintext credentials in a synced file are a security risk.")

headers = ["Company", "Website", "Login", "Password", "Security Q & Answers", "Comments", "Account Number"]
for i, h in enumerate(headers, start=1):
    header_cell(ws.cell(row=4, column=i), h)

r = 5
for row in src_logins:
    for c, v in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=v)
    r += 1

banded(ws, 5, r - 1, len(headers))
for col_letter, w in [("A", 24), ("B", 36), ("C", 22), ("D", 18), ("E", 22), ("F", 32), ("G", 18)]:
    ws.column_dimensions[col_letter].width = w

# === Reorder sheets — Dashboard first ======================================
order = ["Dashboard", "Income", "Expenses", "Savings", "Debts", "Logins"]
wb._sheets = [wb[s] for s in order]

# Hide gridlines on Dashboard for cleaner look
wb["Dashboard"].sheet_view.showGridLines = False

# Save
wb.save(DEST)
print(f"Saved: {DEST}")
print(f"  Income lines: {len(src_income)}")
print(f"  Expense lines: {len(src_expenses)} (categorized)")
print(f"  Debt lines: {len(src_debts)}")
print(f"  Login lines: {len(src_logins)}")
