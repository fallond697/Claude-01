"""
Create a standalone Full Calendar Excel from ServiceNow change data.
Only includes the last 3 months based on Planned Start date.
"""
import json
import os
import calendar
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ──
OUTPUT = os.path.expanduser(
    r"~\OneDrive - Vituity\Documents\Change Management\Change_Management_Calendar.xlsx"
)
SNOW_URL = "https://vituity.service-now.com/nav_to.do?uri=change_request.do?sysparm_query=number="
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "calendar_changes.json")

with open(DATA_FILE, "r", encoding="utf-8") as f:
    all_changes = json.load(f)

# Filter: exclude Canceled/New/Assess, and only last 3 months
EXCLUDE_STATES = {"Canceled", "New", "Assess"}
today = datetime.now().date()
window_start = today - timedelta(days=30)
window_end = today + timedelta(days=14)

changes = []
for c in all_changes:
    if c.get("state", "") in EXCLUDE_STATES:
        continue
    ps = c.get("planned_start", "")
    if ps and len(ps) >= 10:
        try:
            dt = datetime.strptime(ps[:10], "%Y-%m-%d").date()
            if window_start <= dt <= window_end:
                c["_date"] = dt
                changes.append(c)
        except ValueError:
            pass

print(f"Filtered to {len(changes)} changes (last 3 months, excl Canceled/New)")

# ── STYLES ──
TYPE_FILLS = {
    "Normal": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
    "Standard": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    "Emergency": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
}
TYPE_FONTS = {
    "Normal": Font(name="Calibri", size=9, color="FFFFFF"),
    "Standard": Font(name="Calibri", size=9, color="FFFFFF"),
    "Emergency": Font(name="Calibri", size=9, color="FFFFFF", bold=True),
}
DAY_NUM_FONT = Font(name="Calibri", size=12, bold=True, color="333333")
DAY_NUM_FONT_GRAY = Font(name="Calibri", size=12, bold=True, color="CCCCCC")
MONTH_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="003366")
DOW_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
DOW_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
WEEKEND_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
TODAY_BORDER = Border(
    left=Side(style="medium", color="FF6600"),
    right=Side(style="medium", color="FF6600"),
    top=Side(style="medium", color="FF6600"),
    bottom=Side(style="medium", color="FF6600"),
)
CELL_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Build date -> changes lookup
changes_by_date = defaultdict(list)
for c in changes:
    date_key = c.get("planned_start", "")[:10]
    changes_by_date[date_key].append(c)

# Determine months to render
months_to_render = set()
for c in changes:
    ps = c.get("planned_start", "")
    if ps and len(ps) >= 7:
        months_to_render.add(ps[:7])
months_to_render = sorted(months_to_render)

print(f"Months: {', '.join(months_to_render)}")

# ── BUILD WORKBOOK ──
wb = Workbook()
ws = wb.active
ws.title = "Change Calendar"

# Column widths
COL_WIDTH = 24
for col_idx in range(1, 8):
    ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTH

MAX_CHANGES_PER_DAY = 8
DOW_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

current_row = 1

# Title
ws.cell(row=current_row, column=1, value="Change Management Calendar").font = Font(name="Calibri", size=18, bold=True, color="003366")
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
current_row += 1

ws.cell(row=current_row, column=1, value=f"Source: ServiceNow  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {len(changes)} changes").font = Font(name="Calibri", size=9, color="666666")
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
current_row += 1

# Legend
for i, (tname, tfill) in enumerate(TYPE_FILLS.items()):
    col = (i * 2) + 1
    cell = ws.cell(row=current_row, column=col, value=f"  {tname}  ")
    cell.fill = tfill
    cell.font = TYPE_FONTS[tname]
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Count by type for legend
from collections import Counter
type_counts = Counter(c.get("type", "") for c in changes)
col = 7
cell = ws.cell(row=current_row, column=col, value=f"N:{type_counts.get('Normal',0)}  S:{type_counts.get('Standard',0)}  E:{type_counts.get('Emergency',0)}")
cell.font = Font(name="Calibri", size=9, color="333333", bold=True)
cell.alignment = Alignment(horizontal="right")

current_row += 2

# Render each month
for month_str in months_to_render:
    year, month = int(month_str[:4]), int(month_str[5:7])
    month_name = calendar.month_name[month]

    # Count changes this month
    month_count = sum(1 for c in changes if c.get("planned_start", "")[:7] == month_str)

    # Month title
    title_cell = ws.cell(row=current_row, column=1, value=f"{month_name} {year}")
    title_cell.font = MONTH_TITLE_FONT
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    count_cell = ws.cell(row=current_row, column=6, value=f"{month_count} changes")
    count_cell.font = Font(name="Calibri", size=11, color="666666", italic=True)
    count_cell.alignment = Alignment(horizontal="right")
    ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
    current_row += 1

    # Day-of-week headers
    for col_idx, dow in enumerate(DOW_NAMES, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=dow)
        cell.font = DOW_FONT
        cell.fill = DOW_FILL
        cell.alignment = Alignment(horizontal="center")
    current_row += 1

    # Calendar grid (Sunday-start)
    cal = calendar.Calendar(firstweekday=6)
    month_days = cal.monthdayscalendar(year, month)

    for week in month_days:
        week_start_row = current_row

        # Day number row
        for col_idx, day in enumerate(week, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = CELL_BORDER

            if day != 0:
                cell.value = day
                cell.font = DAY_NUM_FONT
                cell.alignment = Alignment(horizontal="left", vertical="top")

                # Count changes for this day
                date_key = f"{year:04d}-{month:02d}-{day:02d}"
                day_count = len(changes_by_date.get(date_key, []))
                if day_count > 0:
                    # Add count badge in same cell
                    cell.value = f"{day}  ({day_count})"

                # Highlight today
                try:
                    this_date = datetime(year, month, day).date()
                    if this_date == today:
                        cell.border = TODAY_BORDER
                        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                except ValueError:
                    pass

                # Weekend shading
                if col_idx in (1, 7) and not (day != 0 and datetime(year, month, day).date() == today):
                    cell.fill = WEEKEND_FILL
            else:
                cell.fill = WEEKEND_FILL

        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Change entry rows
        for entry_idx in range(MAX_CHANGES_PER_DAY):
            has_content = False
            for col_idx, day in enumerate(week, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = CELL_BORDER

                if day == 0:
                    cell.fill = WEEKEND_FILL
                    continue

                date_key = f"{year:04d}-{month:02d}-{day:02d}"
                day_changes = changes_by_date.get(date_key, [])

                if col_idx in (1, 7):
                    cell.fill = WEEKEND_FILL

                # Carry today highlight
                try:
                    if datetime(year, month, day).date() == today:
                        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                except ValueError:
                    pass

                if entry_idx < len(day_changes):
                    has_content = True
                    chg = day_changes[entry_idx]
                    chg_num = chg.get("number", "")
                    short_desc = chg.get("short_description", "")
                    chg_type = chg.get("type", "Normal")
                    state = chg.get("state", "")

                    # Format: CHG#: Description
                    label = f"{chg_num}: {short_desc}"
                    if len(label) > 32:
                        label = label[:30] + ".."

                    cell.value = label
                    cell.fill = TYPE_FILLS.get(chg_type, TYPE_FILLS["Normal"])
                    cell.font = TYPE_FONTS.get(chg_type, TYPE_FONTS["Normal"])
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

                    if chg_num.startswith("CHG"):
                        cell.hyperlink = SNOW_URL + chg_num

                elif entry_idx == MAX_CHANGES_PER_DAY - 1 and len(day_changes) > MAX_CHANGES_PER_DAY:
                    has_content = True
                    remaining = len(day_changes) - MAX_CHANGES_PER_DAY + 1
                    cell.value = f"  +{remaining} more..."
                    cell.font = Font(name="Calibri", size=8, italic=True, color="666666")

            ws.row_dimensions[current_row].height = 14
            current_row += 1

            # If no content was written in this entry row for any day, we can stop early
            # (but keep going to maintain grid alignment)

    # Gap between months
    current_row += 1

# Print area
ws.sheet_properties.pageSetUpPr = None
ws.print_area = f"A1:G{current_row}"

# ── SAVE ──
wb.save(OUTPUT)
print(f"\nCalendar saved to: {OUTPUT}")
print(f"  {len(changes)} changes across {len(months_to_render)} months")
print(f"  Months: {', '.join(months_to_render)}")
