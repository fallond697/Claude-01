#!/usr/bin/env python
"""
Create a Power BI-ready Excel workbook from Change Management SharePoint data.

Source: SharePoint Excel export (JSON tool-results file)
Output: Styled multi-sheet Excel workbook with pivots and dashboard summary.
"""

import json
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- Configuration ---
SOURCE_FILE = (
    r"C:\Users\FallonD\.claude\projects\C--Users-FallonD-Code-Claude-01"
    r"\f07f6e4f-8873-4fff-940c-fe388023aff3\tool-results"
    r"\mcp-sharepoint-read_excel_file-1773075497680.txt"
)
OUTPUT_FILE = (
    r"C:\Users\FallonD\OneDrive - Vituity\Documents\Change Management"
    r"\Change_Management_Dashboard.xlsx"
)
SNOW_URL = "https://vituity.service-now.com/nav_to.do?uri=change_request.do?sysparm_query=number="

# Column indices
COL_NUMBER = 0
COL_SHORT_DESC = 1
COL_DESC = 2
COL_TYPE = 3
COL_STATE = 4
COL_CAB_DATE = 5
COL_PLANNED_START = 6
COL_PLANNED_END = 7
COL_CONFIG_ITEM = 8
COL_ENVIRONMENT = 9
COL_CR_TYPE = 10
COL_REQUESTED_BY = 11
COL_ASSIGNED_TO = 12
COL_ASSIGNMENT_GROUP = 13
COL_CLOSE_CODE = 14
COL_CLOSED_BY = 15
COL_IMPL_PLAN = 16
COL_RISK_IMPACT = 17
COL_BACKOUT_PLAN = 18
COL_TEST_PLAN = 19

DATE_COLS = {COL_CAB_DATE, COL_PLANNED_START, COL_PLANNED_END}
CHANGE_TYPES = ["Normal", "Standard", "Emergency"]

# --- Styles ---
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
PCT_FORMAT = "0.0%"
DATE_FORMAT = "YYYY-MM-DD"


def excel_serial_to_date(serial):
    """Convert Excel serial number to datetime."""
    if serial is None or serial == "" or serial == 0:
        return None
    try:
        serial = float(serial)
        return datetime(1899, 12, 30) + timedelta(days=serial)
    except (ValueError, TypeError):
        return None


def load_data(filepath):
    """Load and parse the JSON tool-results file."""
    with open(filepath, "r", encoding="utf-8") as f:
        wrapper = json.load(f)

    inner = json.loads(wrapper[0]["text"])
    raw_rows = inner["data"]
    headers = raw_rows[0]
    data_rows = raw_rows[1:]
    return headers, data_rows


def style_header_row(ws, num_cols):
    """Apply header styling to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_fit_columns(ws, max_width=50):
    """Estimate and set column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                val_str = str(cell.value)
                # Use first line only for multi-line cells
                first_line = val_str.split("\n")[0] if "\n" in val_str else val_str
                max_len = max(max_len, len(first_line))
        width = min(max_len + 2, max_width)
        ws.column_dimensions[col_letter].width = max(width, 10)


def apply_alt_shading(ws, start_row, end_row, num_cols):
    """Apply alternating row shading."""
    for row_idx in range(start_row, end_row + 1):
        if row_idx % 2 == 0:
            for col in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=col).fill = ALT_FILL


def freeze_top_row(ws):
    """Freeze the top row."""
    ws.freeze_panes = "A2"


def write_pivot_sheet(wb, sheet_name, row_label, pivot_data, types, sort_desc=False,
                      top_n=None):
    """
    Write a pivot table sheet.
    pivot_data: dict of {row_key: {type: count}}
    """
    ws = wb.create_sheet(title=sheet_name)

    # Headers
    headers = [row_label] + types + ["Total"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    # Build rows with totals
    rows = []
    for key, type_counts in pivot_data.items():
        total = sum(type_counts.get(t, 0) for t in types)
        rows.append((key, type_counts, total))

    if sort_desc:
        rows.sort(key=lambda x: x[2], reverse=True)

    if top_n:
        rows = rows[:top_n]

    for r_idx, (key, type_counts, total) in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=key or "(blank)")
        for t_idx, t in enumerate(types, 2):
            ws.cell(row=r_idx, column=t_idx, value=type_counts.get(t, 0))
        ws.cell(row=r_idx, column=len(types) + 2, value=total)

    num_cols = len(headers)
    style_header_row(ws, num_cols)
    apply_alt_shading(ws, 2, ws.max_row, num_cols)
    auto_fit_columns(ws)
    freeze_top_row(ws)
    return ws


def get_month_key(row):
    """Extract YYYY-MM from planned start date."""
    dt = excel_serial_to_date(row[COL_PLANNED_START])
    if dt:
        return dt.strftime("%Y-%m")
    # Fallback to CAB date
    dt = excel_serial_to_date(row[COL_CAB_DATE])
    if dt:
        return dt.strftime("%Y-%m")
    return "(no date)"


def build_pivot(rows, key_col_fn, type_col=COL_TYPE):
    """Build a pivot dict: {key: {type: count}}."""
    pivot = defaultdict(lambda: defaultdict(int))
    for row in rows:
        key = key_col_fn(row)
        change_type = row[type_col] if row[type_col] else "(blank)"
        pivot[key][change_type] += 1
    return pivot


def main():
    print("Loading source data...")
    headers, data_rows = load_data(SOURCE_FILE)
    print(f"  Loaded {len(data_rows)} change records with {len(headers)} columns")

    wb = Workbook()

    # =========================================================================
    # Sheet 1: Raw Data
    # =========================================================================
    print("Creating Raw Data sheet...")
    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    # Write headers
    for c, h in enumerate(headers, 1):
        ws_raw.cell(row=1, column=c, value=h)

    # Write data
    for r_idx, row in enumerate(data_rows, 2):
        for c_idx, val in enumerate(row):
            cell = ws_raw.cell(row=r_idx, column=c_idx + 1)

            if c_idx in DATE_COLS:
                dt = excel_serial_to_date(val)
                if dt:
                    cell.value = dt
                    cell.number_format = "YYYY-MM-DD"
                else:
                    cell.value = val
            elif c_idx == COL_NUMBER and val and str(val).startswith("CHG"):
                cell.value = val
                cell.hyperlink = SNOW_URL + str(val)
                cell.font = Font(color="0563C1", underline="single")
            else:
                cell.value = val

            # Wrap long text columns
            if c_idx in (COL_DESC, COL_IMPL_PLAN, COL_RISK_IMPACT,
                         COL_BACKOUT_PLAN, COL_TEST_PLAN):
                cell.alignment = WRAP_ALIGNMENT

    num_cols = len(headers)
    style_header_row(ws_raw, num_cols)
    apply_alt_shading(ws_raw, 2, ws_raw.max_row, num_cols)
    auto_fit_columns(ws_raw)
    freeze_top_row(ws_raw)

    # =========================================================================
    # Sheet 2: Changes by Month
    # =========================================================================
    print("Creating Changes by Month sheet...")
    month_pivot = build_pivot(data_rows, get_month_key)
    # Sort months chronologically
    sorted_months = dict(sorted(month_pivot.items()))
    write_pivot_sheet(wb, "Changes by Month", "Month", sorted_months, CHANGE_TYPES)

    # =========================================================================
    # Sheet 3: Changes by Config Item
    # =========================================================================
    print("Creating Changes by Config Item sheet...")
    ci_pivot = build_pivot(data_rows, lambda r: r[COL_CONFIG_ITEM] or "(blank)")
    write_pivot_sheet(wb, "Changes by Config Item", "Configuration Item", ci_pivot,
                      CHANGE_TYPES, sort_desc=True, top_n=20)

    # =========================================================================
    # Sheet 4: Changes by State
    # =========================================================================
    print("Creating Changes by State sheet...")
    state_pivot = build_pivot(data_rows, lambda r: r[COL_STATE] or "(blank)")
    write_pivot_sheet(wb, "Changes by State", "State", state_pivot, CHANGE_TYPES)

    # =========================================================================
    # Sheet 5: Changes by Assignment Group
    # =========================================================================
    print("Creating Changes by Assignment Group sheet...")
    ag_pivot = build_pivot(data_rows, lambda r: r[COL_ASSIGNMENT_GROUP] or "(blank)")
    write_pivot_sheet(wb, "Changes by Assignment Group", "Assignment Group", ag_pivot,
                      CHANGE_TYPES, sort_desc=True, top_n=15)

    # =========================================================================
    # Sheet 6: Changes by Environment
    # =========================================================================
    print("Creating Changes by Environment sheet...")
    env_pivot = build_pivot(data_rows, lambda r: r[COL_ENVIRONMENT] or "(blank)")
    write_pivot_sheet(wb, "Changes by Environment", "Environment", env_pivot,
                      CHANGE_TYPES)

    # =========================================================================
    # Sheet 7: Monthly Trend
    # =========================================================================
    print("Creating Monthly Trend sheet...")
    ws_trend = wb.create_sheet(title="Monthly Trend")
    trend_headers = ["Month", "Total Changes", "Normal", "Standard", "Emergency",
                     "Closed", "Success Rate"]
    for c, h in enumerate(trend_headers, 1):
        ws_trend.cell(row=1, column=c, value=h)

    # Build monthly stats
    monthly_total = defaultdict(int)
    monthly_type = defaultdict(lambda: defaultdict(int))
    monthly_closed = defaultdict(int)
    monthly_successful = defaultdict(int)

    for row in data_rows:
        month = get_month_key(row)
        monthly_total[month] += 1
        change_type = row[COL_TYPE] or "(blank)"
        monthly_type[month][change_type] += 1
        if row[COL_STATE] and row[COL_STATE].lower() == "closed":
            monthly_closed[month] += 1
            close_code = (row[COL_CLOSE_CODE] or "").lower()
            if "successful" in close_code or "success" in close_code:
                monthly_successful[month] += 1

    sorted_months_list = sorted(monthly_total.keys())
    for r_idx, month in enumerate(sorted_months_list, 2):
        ws_trend.cell(row=r_idx, column=1, value=month)
        ws_trend.cell(row=r_idx, column=2, value=monthly_total[month])
        ws_trend.cell(row=r_idx, column=3, value=monthly_type[month].get("Normal", 0))
        ws_trend.cell(row=r_idx, column=4, value=monthly_type[month].get("Standard", 0))
        ws_trend.cell(row=r_idx, column=5,
                      value=monthly_type[month].get("Emergency", 0))
        closed = monthly_closed[month]
        ws_trend.cell(row=r_idx, column=6, value=closed)
        if closed > 0:
            rate = monthly_successful[month] / closed
            cell = ws_trend.cell(row=r_idx, column=7, value=rate)
            cell.number_format = PCT_FORMAT
        else:
            ws_trend.cell(row=r_idx, column=7, value="N/A")

    style_header_row(ws_trend, len(trend_headers))
    apply_alt_shading(ws_trend, 2, ws_trend.max_row, len(trend_headers))
    auto_fit_columns(ws_trend)
    freeze_top_row(ws_trend)

    # =========================================================================
    # Sheet 8: Dashboard Summary
    # =========================================================================
    print("Creating Dashboard Summary sheet...")
    ws_dash = wb.create_sheet(title="Dashboard Summary")

    # Gather stats
    total_changes = len(data_rows)
    type_counts = Counter(r[COL_TYPE] or "(blank)" for r in data_rows)
    state_counts = Counter(r[COL_STATE] or "(blank)" for r in data_rows)
    ci_counts = Counter(r[COL_CONFIG_ITEM] or "(blank)" for r in data_rows)
    ag_counts = Counter(r[COL_ASSIGNMENT_GROUP] or "(blank)" for r in data_rows)

    num_months = len(set(get_month_key(r) for r in data_rows) - {"(no date)"})
    avg_per_month = total_changes / num_months if num_months > 0 else 0

    total_closed = sum(1 for r in data_rows
                       if r[COL_STATE] and r[COL_STATE].lower() == "closed")
    total_successful = sum(
        1 for r in data_rows
        if r[COL_STATE] and r[COL_STATE].lower() == "closed"
        and ("successful" in (r[COL_CLOSE_CODE] or "").lower()
             or "success" in (r[COL_CLOSE_CODE] or "").lower())
    )
    success_rate = (total_successful / total_closed * 100) if total_closed > 0 else 0

    section_font = Font(bold=True, size=13, color="003366")
    metric_font = Font(size=11)
    value_font = Font(bold=True, size=11)

    row = 1

    def write_section(title):
        nonlocal row
        ws_dash.cell(row=row, column=1, value=title).font = section_font
        row += 1

    def write_metric(label, value, fmt=None):
        nonlocal row
        ws_dash.cell(row=row, column=1, value=label).font = metric_font
        cell = ws_dash.cell(row=row, column=2, value=value)
        cell.font = value_font
        if fmt:
            cell.number_format = fmt
        row += 1

    write_section("Overall Metrics")
    write_metric("Total Changes", total_changes)
    write_metric("Months Covered", num_months)
    write_metric("Average Changes/Month", round(avg_per_month, 1))
    write_metric("Overall Success Rate", f"{success_rate:.1f}%")
    row += 1

    write_section("Changes by Type")
    for t in CHANGE_TYPES:
        write_metric(t, type_counts.get(t, 0))
    for t in sorted(type_counts.keys()):
        if t not in CHANGE_TYPES:
            write_metric(t, type_counts[t])
    row += 1

    write_section("Changes by State")
    for state, count in state_counts.most_common():
        write_metric(state, count)
    row += 1

    write_section("Top 5 Configuration Items")
    for ci, count in ci_counts.most_common(5):
        write_metric(ci, count)
    row += 1

    write_section("Top 5 Assignment Groups")
    for ag, count in ag_counts.most_common(5):
        write_metric(ag, count)

    # Style dashboard
    ws_dash.column_dimensions["A"].width = 40
    ws_dash.column_dimensions["B"].width = 20
    freeze_top_row(ws_dash)

    # =========================================================================
    # Save
    # =========================================================================
    print(f"\nSaving workbook to:\n  {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)

    # Summary
    print("\n" + "=" * 60)
    print("DASHBOARD CREATION COMPLETE")
    print("=" * 60)
    print(f"  Records processed:    {total_changes}")
    print(f"  Months covered:       {num_months}")
    print(f"  Sheets created:       {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"    - {name}")
    print(f"\n  By Type:")
    for t in CHANGE_TYPES:
        print(f"    {t:15s} {type_counts.get(t, 0):>5}")
    print(f"\n  By State:")
    for s, c in state_counts.most_common():
        print(f"    {s:15s} {c:>5}")
    print(f"\n  Success Rate:         {success_rate:.1f}%")
    print(f"  Avg Changes/Month:   {avg_per_month:.1f}")
    print(f"\n  Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
