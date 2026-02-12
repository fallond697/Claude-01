#!/usr/bin/env python3
"""Format the 2026 Project Timeline sheet with Gantt-style visual bars."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import shutil
import sys

SRC = r'C:\Users\FallonD\OneDrive - Vituity\2026 Product Project Plan for Change Management.xlsx'
TMP = r'C:\Users\FallonD\AppData\Local\Temp\2026_timeline_formatted.xlsx'

# =====================================================================
# Color palette
# =====================================================================
PHASE_COLORS = {
    'Planning':    PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    'Development': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'Execution':   PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'Monitoring':  PatternFill(start_color='E2D0F0', end_color='E2D0F0', fill_type='solid'),
    'Initiation':  PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
    'Completed':   PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
}

PHASE_FONT_COLORS = {
    'Planning':    Font(color='1F4E79', size=9, bold=True),
    'Development': Font(color='006100', size=9, bold=True),
    'Execution':   Font(color='BF4000', size=9, bold=True),
    'Monitoring':  Font(color='6B3FA0', size=9, bold=True),
    'Initiation':  Font(color='7F6000', size=9, bold=True),
    'Completed':   Font(color='595959', size=9, bold=True),
}

BLOCKER_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', size=10, bold=True)
TITLE_FONT = Font(size=14, bold=True, color='2F5496')
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
MONTH_EMPTY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

# Month column mapping: J=10 (Jan) through U=21 (Dec)
MONTH_COLS = {1: 10, 2: 11, 3: 12, 4: 13, 5: 14, 6: 15,
              7: 16, 8: 17, 9: 18, 10: 19, 11: 20, 12: 21}


def get_active_months(start_val, end_val):
    """Return list of month numbers (1-12) that are active."""
    start_date = None
    end_date = None

    for val, target in [(start_val, 'start'), (end_val, 'end')]:
        parsed = None
        if isinstance(val, datetime):
            parsed = val
        elif isinstance(val, str) and val.strip():
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y']:
                try:
                    parsed = datetime.strptime(val.strip(), fmt)
                    break
                except ValueError:
                    pass
        if target == 'start':
            start_date = parsed
        else:
            end_date = parsed

    if not start_date:
        return None  # signal to use existing cell values

    if not end_date:
        end_date = datetime(2026, 12, 31)

    # Clamp to 2026
    if start_date.year > 2026 or end_date.year < 2026:
        return []

    start_m = start_date.month if start_date.year == 2026 else 1
    end_m = end_date.month if end_date.year == 2026 else 12

    if start_date.year < 2026:
        start_m = 1

    return list(range(max(1, start_m), min(12, end_m) + 1))


def format_timeline():
    # Work on temp copy
    shutil.copy2(SRC, TMP)
    wb = openpyxl.load_workbook(TMP)
    tl = wb['2026 Project TImeline']

    # -----------------------------------------------------------------
    # 1. Style header row (row 4)
    # -----------------------------------------------------------------
    for col in range(1, 22):
        cell = tl.cell(4, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Title
    tl.cell(1, 1).font = TITLE_FONT

    # -----------------------------------------------------------------
    # 2. Legend in row 3
    # -----------------------------------------------------------------
    legend_row = 3
    tl.cell(legend_row, 1).value = 'Legend:'
    tl.cell(legend_row, 1).font = Font(size=9, bold=True)
    tl.cell(legend_row, 1).alignment = Alignment(vertical='center')

    legend_items = [
        ('Planning', 2), ('Development', 3), ('Execution', 4),
        ('Monitoring', 5), ('Initiation', 6), ('Completed', 7),
        ('Blocked', 8),
    ]
    for phase, col_idx in legend_items:
        cell = tl.cell(legend_row, col_idx)
        if phase == 'Blocked':
            cell.fill = BLOCKER_FILL
            cell.font = Font(size=9, bold=True, color='9C0006')
        else:
            cell.fill = PHASE_COLORS[phase]
            cell.font = PHASE_FONT_COLORS[phase]
        cell.value = phase
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    for col in range(9, 22):
        tl.cell(legend_row, col).value = None

    # -----------------------------------------------------------------
    # 3. Format data rows with Gantt bars
    # -----------------------------------------------------------------
    data_start_row = 5
    data_end_row = tl.max_row

    for row in range(data_start_row, data_end_row + 1):
        app_name = tl.cell(row, 1).value
        if not app_name:
            continue

        phase = str(tl.cell(row, 2).value or 'Planning')
        start_val = tl.cell(row, 3).value
        end_val = tl.cell(row, 4).value
        blocker = str(tl.cell(row, 8).value or '').strip().lower()
        is_blocked = blocker in ('yes', 'true')

        phase_fill = PHASE_COLORS.get(phase, PHASE_COLORS['Planning'])
        phase_font = PHASE_FONT_COLORS.get(phase, PHASE_FONT_COLORS['Planning'])

        # Determine active months
        active_months = get_active_months(start_val, end_val)
        if active_months is None:
            active_months = []
            for m, c in MONTH_COLS.items():
                if tl.cell(row, c).value:
                    active_months.append(m)

        # Apply Gantt bars to month columns
        for month_num, col_idx in MONTH_COLS.items():
            cell = tl.cell(row, col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if month_num in active_months:
                if is_blocked:
                    # Striped effect: blocked phase color with red text
                    cell.fill = phase_fill
                    cell.font = Font(color='CC0000', size=11, bold=True)
                    cell.value = '\u2588'
                else:
                    cell.fill = phase_fill
                    cell.font = phase_font
                    cell.value = '\u2588'
            else:
                cell.fill = MONTH_EMPTY_FILL
                cell.font = Font(color='D9D9D9', size=9)
                cell.value = None

        # Format data columns A-I
        for col in range(1, 10):
            cell = tl.cell(row, col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            if is_blocked and col == 1:
                cell.fill = BLOCKER_FILL
                cell.font = Font(bold=True, color='9C0006', size=10)
            elif col == 1:
                cell.font = Font(bold=True, size=10)
            elif col == 8 and is_blocked:
                cell.fill = BLOCKER_FILL
                cell.font = Font(bold=True, color='9C0006')

        # Phase column (B) gets phase color
        phase_cell = tl.cell(row, 2)
        phase_cell.fill = phase_fill
        phase_cell.font = phase_font
        phase_cell.alignment = Alignment(horizontal='center', vertical='center')

    # -----------------------------------------------------------------
    # 4. Column widths
    # -----------------------------------------------------------------
    col_widths = {
        'A': 18, 'B': 14, 'C': 12, 'D': 12, 'E': 30,
        'F': 14, 'G': 20, 'H': 9, 'I': 25,
    }
    for letter, width in col_widths.items():
        tl.column_dimensions[letter].width = width

    for month_num, col_idx in MONTH_COLS.items():
        tl.column_dimensions[get_column_letter(col_idx)].width = 5

    # -----------------------------------------------------------------
    # 5. Row heights
    # -----------------------------------------------------------------
    tl.row_dimensions[1].height = 24
    tl.row_dimensions[2].height = 15
    tl.row_dimensions[3].height = 22
    tl.row_dimensions[4].height = 30

    for row in range(data_start_row, data_end_row + 1):
        tl.row_dimensions[row].height = 45

    # -----------------------------------------------------------------
    # 6. Freeze panes
    # -----------------------------------------------------------------
    tl.freeze_panes = 'B5'

    # -----------------------------------------------------------------
    # 7. Month header names - short
    # -----------------------------------------------------------------
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    for i, (month_num, col_idx) in enumerate(MONTH_COLS.items()):
        cell = tl.cell(4, col_idx)
        cell.value = month_names[i]

    # -----------------------------------------------------------------
    # 8. Subtle instruction row
    # -----------------------------------------------------------------
    tl.cell(2, 1).font = Font(size=8, italic=True, color='808080')

    # -----------------------------------------------------------------
    # Save and copy back
    # -----------------------------------------------------------------
    wb.save(TMP)
    shutil.copy2(TMP, SRC)
    print("Timeline formatted successfully!")
    print(f"Saved to: {SRC}")


if __name__ == '__main__':
    format_timeline()
